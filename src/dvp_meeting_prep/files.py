from __future__ import annotations

from csv import DictReader
from datetime import datetime
from pathlib import Path
import json
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _normalize_header(header: Any) -> str:
    if header is None:
        return ""
    text = str(header).strip().lower()
    text = text.replace("\n", " ")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    text = str(value).strip()
    return text if text != "" else None


def _to_number(value: Any) -> str | None:
    cleaned = _clean_value(value)
    if cleaned is None:
        return None
    return str(cleaned)


def _to_float(value: Any) -> float | None:
    cleaned = _clean_value(value)
    if cleaned is None:
        return None
    try:
        return float(str(cleaned).replace(",", ""))
    except ValueError:
        return None


def _to_int(value: Any) -> int | None:
    number = _to_float(value)
    if number is None:
        return None
    return int(number)


def _to_bool(value: Any) -> bool | None:
    cleaned = _clean_value(value)
    if cleaned is None:
        return None
    lowered = str(cleaned).strip().lower()
    if lowered in {"yes", "y", "true", "1"}:
        return True
    if lowered in {"no", "n", "false", "0"}:
        return False
    return None


def _to_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    cleaned = _clean_value(value)
    if cleaned is None:
        return None
    text = str(cleaned)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%B %d, %Y", "%b %d, %Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _is_blank_row(row: list[Any]) -> bool:
    return all(_clean_value(value) is None for value in row)


def _find_header_row(rows: list[list[Any]], expected_fields: set[str]) -> tuple[int, list[Any]]:
    expected_normalized = {_normalize_header(value) for value in expected_fields}
    for index, row in enumerate(rows):
        normalized = {_normalize_header(value) for value in row if _clean_value(value) is not None}
        if expected_normalized.issubset(normalized):
            return index, row
    raise RuntimeError(f"Could not locate a header row containing {sorted(expected_normalized)}")


def _find_value_by_header(headers: list[Any], row: list[Any], target_header: str, *, exact: bool = False) -> Any:
    target = _normalize_header(target_header)
    for index, header in enumerate(headers):
        if exact:
            if _clean_value(header) is not None and str(header).strip().lower() == target_header.strip().lower():
                return _clean_value(row[index] if index < len(row) else None)
            continue
        if _normalize_header(header) == target:
            return _clean_value(row[index] if index < len(row) else None)
    return None


def _row_payload(headers: list[Any], row: list[Any]) -> list[dict[str, Any]]:
    payload: list[dict[str, Any]] = []
    for index, header in enumerate(headers):
        value = row[index] if index < len(row) else None
        if _clean_value(header) is None and _clean_value(value) is None:
            continue
        payload.append(
            {
                "column_index": index + 1,
                "header": _clean_value(header),
                "value": _clean_value(value),
            }
        )
    return payload


def _forward_fill_headers(values: list[Any]) -> list[str | None]:
    filled: list[str | None] = []
    last: str | None = None
    for value in values:
        cleaned = _clean_value(value)
        if cleaned is not None:
            last = str(cleaned)
        filled.append(last)
    return filled


def _dedup_parts(parts: list[str]) -> list[str]:
    deduped: list[str] = []
    for part in parts:
        if not deduped or deduped[-1] != part:
            deduped.append(part)
    return deduped


def _build_consultant_header_map(rows: list[list[Any]], header_index: int, headers_en: list[Any]) -> list[dict[str, Any]]:
    row1 = list(rows[0]) if len(rows) > 0 else []
    row2 = list(rows[1]) if len(rows) > 1 else []
    row4 = list(rows[header_index + 1]) if len(rows) > header_index + 1 else []

    size = len(headers_en)
    while len(row1) < size:
        row1.append(None)
    while len(row2) < size:
        row2.append(None)
    while len(row4) < size:
        row4.append(None)

    ff1 = _forward_fill_headers(row1)
    ff2 = _forward_fill_headers(row2)

    static_headers = {
        _normalize_header("Advisor#"),
        _normalize_header("Advisor"),
        _normalize_header("Area"),
        _normalize_header("RO#"),
        _normalize_header("Region"),
        _normalize_header("Div"),
        _normalize_header("Base Achievement Level"),
        _normalize_header("ETF Completed & Approved"),
        _normalize_header("Designation"),
        _normalize_header("Sales Start Date"),
        _normalize_header("Termination Date"),
        _normalize_header("Tenure Category"),
        _normalize_header("PWM Indicator"),
        _normalize_header("Dealer Code"),
        _normalize_header("Insurance Expiry Date"),
    }

    header_map: list[dict[str, Any]] = []
    for i in range(size):
        h_en = _clean_value(headers_en[i] if i < len(headers_en) else None)
        h_fr = _clean_value(row4[i] if i < len(row4) else None)
        g1 = _clean_value(ff1[i] if i < len(ff1) else None)
        g2 = _clean_value(ff2[i] if i < len(ff2) else None)

        g1_norm = _normalize_header(g1)
        if g1_norm == _normalize_header("Monthly Advisor Detail Report / Rapport détaillé des conseillers mensuel"):
            g1 = None

        leaf_norm = _normalize_header(h_en)
        is_metric = bool(h_en) and leaf_norm not in static_headers

        parts = [part for part in [g1, g2, h_en] if part]
        parts = _dedup_parts(parts)
        metric_key = _normalize_header("__".join(parts)) if is_metric else None

        header_map.append(
            {
                "column_index": i + 1,
                "group_level_1": g1,
                "group_level_2": g2,
                "header_en": h_en,
                "header_fr": h_fr,
                "is_metric": is_metric,
                "metric_key": metric_key,
                "metric_label": " > ".join(parts) if is_metric else h_en,
            }
        )

    return header_map


SCORECARD_LAST_COL_INDEX = 84  # A:CF


def _scorecard_group_from_column(column_index: int) -> str:
    if 16 <= column_index <= 35:
        return "Net Contributions"
    if 36 <= column_index <= 41:
        return "New Business"
    if 42 <= column_index <= 45:
        return "New Business Detail"
    if 46 <= column_index <= 51:
        return "First Year Commission Insurance"
    if 52 <= column_index <= 58:
        return "Mortgage / HEP"
    if column_index == 59:
        return "Key Driver Score"
    if 60 <= column_index <= 66:
        return "Client"
    if 67 <= column_index <= 69:
        return "Growth"
    if column_index == 70:
        return "Family Group"
    if 71 <= column_index <= 84:
        return "Assets"
    return "Advisor Profile"


def _normalize_period_label(label: Any) -> str | None:
    cleaned = _clean_value(label)
    if cleaned is None:
        return None
    text = _normalize_header(cleaned)
    if "rolling_12" in text or "12_derniers_mois" in text:
        return "Rolling 12"
    if "month" in text or "mois" in text:
        return "Month"
    if "ytd" in text or "dda" in text:
        return "YTD"
    return str(cleaned)


def _safe_numeric(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = _clean_value(value)
    if cleaned is None:
        return None
    text = str(cleaned).replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def _parse_report_date(raw: Any) -> str:
    cleaned = _clean_value(raw)
    if cleaned is None:
        raise RuntimeError("Could not parse consultant scorecard report date from A2.")
    head = str(cleaned).split("/")[0].strip()
    parsed = _to_date(head)
    if parsed is None:
        raise RuntimeError(f"Could not parse consultant scorecard report date from value: {cleaned}")
    return parsed


def parse_consultant_scorecard(path: str | Path, *, source_file_name: str | None = None) -> dict[str, Any]:
    """Parse a consultant scorecard workbook.

    `source_file_name` overrides the recorded source_file value. This matters
    for uploads: the on-disk path is a randomly named temp file, and using
    that name would make the same logical file hash differently on every
    upload, defeating dedup.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return _parse_consultant_scorecard_workbook(workbook, source_file_name or Path(path).name)
    finally:
        # read_only workbooks keep the underlying zip file handle open until
        # closed, which on Windows blocks deleting/replacing the source file.
        workbook.close()


def _parse_consultant_scorecard_workbook(workbook: Any, source_file_name: str) -> dict[str, Any]:
    sheet = None
    for candidate in workbook.worksheets:
        if "advisor detail" in candidate.title.lower():
            sheet = candidate
            break
    if sheet is None:
        raise RuntimeError("Could not find consultant scorecard worksheet containing 'Advisor Detail'.")

    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        return {
            "sheet_name": sheet.title,
            "report_date": None,
            "header_rows": {"major": 1, "period": 2, "english": 3, "french": 4},
            "raw_rows": [],
            "monthly_rows": [],
            "metric_rows": [],
        }

    rows = [list(row[:SCORECARD_LAST_COL_INDEX]) for row in all_rows]
    report_date = _parse_report_date(rows[1][0] if len(rows) > 1 and len(rows[1]) > 0 else None)
    headers_major = rows[0] if len(rows) > 0 else []
    headers_period = rows[1] if len(rows) > 1 else []
    headers_english = rows[2] if len(rows) > 2 else []
    headers_french = rows[3] if len(rows) > 3 else []

    while len(headers_major) < SCORECARD_LAST_COL_INDEX:
        headers_major.append(None)
    while len(headers_period) < SCORECARD_LAST_COL_INDEX:
        headers_period.append(None)
    while len(headers_english) < SCORECARD_LAST_COL_INDEX:
        headers_english.append(None)
    while len(headers_french) < SCORECARD_LAST_COL_INDEX:
        headers_french.append(None)

    ff_period = _forward_fill_headers(headers_period)

    raw_rows: list[dict[str, Any]] = []
    monthly_rows: list[dict[str, Any]] = []
    metric_rows: list[dict[str, Any]] = []

    for row_number, row in enumerate(rows[4:], start=5):
        if len(row) < SCORECARD_LAST_COL_INDEX:
            row = row + [None] * (SCORECARD_LAST_COL_INDEX - len(row))

        advisor_number = _to_number(row[0])
        if advisor_number is None:
            break

        advisor_name = _clean_value(row[1])
        if advisor_name is None:
            continue

        monthly = {
            "source_file": source_file_name,
            "report_date": report_date,
            "source_row_number": row_number,
            "advisor_number": advisor_number,
            "advisor_name": advisor_name,
            "area": _clean_value(row[2]),
            "ro_number": _to_int(row[3]),
            "region": _clean_value(row[4]),
            "division": _to_int(row[5]),
            "base_achievement_level": _clean_value(row[6]),
            "etf_completed_approved": _to_bool(row[7]),
            "designation": _clean_value(row[8]),
            "sales_start_date": _to_date(row[9]),
            "termination_date": _to_date(row[10]),
            "tenure_category": _clean_value(row[11]),
            "pwm_indicator": _to_bool(row[12]),
            "dealer_code": _clean_value(row[13]),
            "insurance_expiry_date": _to_date(row[14]),
            "key_driver_score": _safe_numeric(row[58]),
            "raw_payload": None,
        }

        row_payload: dict[str, Any] = {}
        metrics_for_payload: list[dict[str, Any]] = []

        for col_idx in range(1, SCORECARD_LAST_COL_INDEX + 1):
            value = row[col_idx - 1]
            cleaned = _clean_value(value)
            if cleaned is None:
                continue

            group_name = _scorecard_group_from_column(col_idx)
            period_name = _normalize_period_label(ff_period[col_idx - 1])
            header_en = _clean_value(headers_english[col_idx - 1])
            header_fr = _clean_value(headers_french[col_idx - 1])
            metric_name = header_en or f"Column {get_column_letter(col_idx)}"

            if col_idx <= 15:
                key = _normalize_header(metric_name)
                if key:
                    row_payload[key] = cleaned
                continue

            value_numeric = _safe_numeric(value)
            value_date = _to_date(value)
            value_text = None if value_numeric is not None or value_date is not None else cleaned
            unit = "percent" if "%" in (metric_name or "") else None

            metric_row = {
                "advisor_number": advisor_number,
                "source_row_number": row_number,
                "source_column": get_column_letter(col_idx),
                "metric_group": group_name,
                "metric_period": period_name,
                "metric_name": metric_name,
                "value_numeric": value_numeric,
                "value_text": value_text,
                "value_date": value_date,
                "unit": unit,
                "french_label": header_fr,
            }
            metric_rows.append(metric_row)
            metrics_for_payload.append(metric_row)

        raw_payload = {
            "report_date": report_date,
            "advisor": {
                "advisor_number": monthly["advisor_number"],
                "advisor_name": monthly["advisor_name"],
                "area": monthly["area"],
                "ro_number": monthly["ro_number"],
                "region": monthly["region"],
                "division": monthly["division"],
                "base_achievement_level": monthly["base_achievement_level"],
                "etf_completed_approved": monthly["etf_completed_approved"],
                "designation": monthly["designation"],
                "sales_start_date": monthly["sales_start_date"],
                "termination_date": monthly["termination_date"],
                "tenure_category": monthly["tenure_category"],
                "pwm_indicator": monthly["pwm_indicator"],
                "dealer_code": monthly["dealer_code"],
                "insurance_expiry_date": monthly["insurance_expiry_date"],
            },
            "profile_values": row_payload,
            "metrics": metrics_for_payload,
        }

        raw_rows.append(
            {
                "source_file": source_file_name,
                "sheet_name": sheet.title,
                "report_date": report_date,
                "source_row_number": row_number,
                "advisor_number": advisor_number,
                "advisor_name": advisor_name,
                "raw_payload": raw_payload,
            }
        )

        monthly["raw_payload"] = raw_payload
        monthly_rows.append(monthly)

    return {
        "sheet_name": sheet.title,
        "report_date": report_date,
        "header_rows": {"major": 1, "period": 2, "english": 3, "french": 4},
        "raw_rows": raw_rows,
        "monthly_rows": monthly_rows,
        "metric_rows": metric_rows,
    }


def read_salesforce_rows(path: str | Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return _read_salesforce_workbook(workbook)
    finally:
        workbook.close()


def _read_salesforce_workbook(workbook: Any) -> list[dict[str, Any]]:
    sheet = None
    for candidate in workbook.worksheets:
        if "ws team crm history" in candidate.title.lower():
            sheet = candidate
            break
    if sheet is None:
        raise RuntimeError("Could not find the Salesforce worksheet named 'WS Team CRM History'.")

    rows = list(sheet.iter_rows(values_only=True))
    header_index, headers = _find_header_row(rows, {"advisor_number", "name", "task_subtype"})

    parsed: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if _is_blank_row(list(row)):
            continue

        raw_payload = {str(headers[i]): _clean_value(row[i] if i < len(row) else None) for i in range(len(headers))}
        advisor_name = _find_value_by_header(headers, list(row), "Name")
        if advisor_name is None:
            continue

        parsed.append(
            {
                "advisor_name": advisor_name,
                "advisor_number": _to_number(_find_value_by_header(headers, list(row), "Advisor Number")),
                "task_subtype": _find_value_by_header(headers, list(row), "Task Subtype"),
                "subject": _find_value_by_header(headers, list(row), "Subject"),
                "comments": _find_value_by_header(headers, list(row), "Comments"),
                "interaction_type": _find_value_by_header(headers, list(row), "Interaction Type"),
                "completed_date_time": _find_value_by_header(headers, list(row), "Completed Date/Time"),
                "district_vp_wholesaling": _find_value_by_header(headers, list(row), "District VP (Wholesaling)"),
                "pwm": _find_value_by_header(headers, list(row), "PWM"),
                "book_size": _find_value_by_header(headers, list(row), "Book Size"),
                "assets_under_management": _find_value_by_header(headers, list(row), "Assets Under Management (AUM)"),
                "new_business_ytd": _find_value_by_header(headers, list(row), "New Business YTD"),
                "created_date": _find_value_by_header(headers, list(row), "Created Date"),
                "start_date": _find_value_by_header(headers, list(row), "Start Date"),
                "status": _find_value_by_header(headers, list(row), "Status"),
                "area": _find_value_by_header(headers, list(row), "Area"),
                "region_office_number": _find_value_by_header(headers, list(row), "Region Office Number"),
                "assigned": _find_value_by_header(headers, list(row), "Assigned"),
                "raw_payload": raw_payload,
            }
        )

    return parsed


def read_tableau_rows(path: str | Path) -> list[dict[str, Any]]:
    parsed: list[dict[str, Any]] = []
    with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
        reader = DictReader(handle)
        for row in reader:
            advisor_name = _clean_value(row.get("Advisor"))
            if advisor_name is None:
                continue

            date_value = _clean_value(row.get("Date"))
            if date_value is not None:
                try:
                    date_value = datetime.strptime(date_value, "%m/%d/%Y").date().isoformat()
                except ValueError:
                    pass

            parsed.append(
                {
                    "advisor_name": advisor_name,
                    "advisor_name_number": _clean_value(row.get("Advisor Name - Number")),
                    "segment": _clean_value(row.get("Segment")),
                    "date": date_value,
                    "area_name": _clean_value(row.get("Area Name")),
                    "region": _clean_value(row.get("Region")),
                    "measure_names": _clean_value(row.get("Measure Names")),
                    "account_count_fund_formatted": _to_float(row.get("Account Count Fund Formatted")),
                    "client_count_fund_formatted": _to_float(row.get("Client Count Fund Formatted")),
                    "fund_formatted": _clean_value(row.get("Fund Formatted")),
                    "approved_to_buy": _clean_value(row.get("Approved to Buy")),
                    "area": _clean_value(row.get("Area")),
                    "division_manager": _clean_value(row.get("Division Manager")),
                    "fund_family": _clean_value(row.get("Fund Family")),
                    "investment_vehicle": _clean_value(row.get("Investment Vehicle")),
                    "pwm": _clean_value(row.get("PWM")),
                    "region_name": _clean_value(row.get("Region Name")),
                    "measure_values": _to_float(row.get("Measure Values")),
                    "raw_payload": {key: _clean_value(value) for key, value in row.items()},
                }
            )

    return parsed


def read_consultant_scorecard_rows(path: str | Path) -> list[dict[str, Any]]:
    structured = parse_consultant_scorecard(path)
    parsed: list[dict[str, Any]] = []
    for row in structured["monthly_rows"]:
        parsed.append(
            {
                "advisor_name": row.get("advisor_name"),
                "advisor_number": row.get("advisor_number"),
                "area": row.get("area"),
                "ro_number": row.get("ro_number"),
                "region": row.get("region"),
                "division": row.get("division"),
                "base_achievement_level": row.get("base_achievement_level"),
                "etf_completed_and_approved": row.get("etf_completed_approved"),
                "designation": row.get("designation"),
                "sales_start_date": row.get("sales_start_date"),
                "termination_date": row.get("termination_date"),
                "tenure_category": row.get("tenure_category"),
                "pwm_indicator": row.get("pwm_indicator"),
                "dealer_code": row.get("dealer_code"),
                "insurance_expiry_date": row.get("insurance_expiry_date"),
                "key_driver_score": row.get("key_driver_score"),
                "client_bp_count": None,
                "assets_under_management": None,
                "third_party_assets": None,
                "assets_under_administration": None,
                "raw_payload": row.get("raw_payload"),
            }
        )
    return parsed


def pretty_json(value: Any) -> str:
    return json.dumps(value, indent=2, ensure_ascii=False, default=str)


def slugify_filename(value: str) -> str:
    slug: list[str] = []
    previous_was_separator = False
    for char in value.lower():
        if char.isalnum():
            slug.append(char)
            previous_was_separator = False
        elif not previous_was_separator:
            slug.append("_")
            previous_was_separator = True
    text = "".join(slug).strip("_")
    return text or "advisor"

